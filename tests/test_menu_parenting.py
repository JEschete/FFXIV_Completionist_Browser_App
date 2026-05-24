from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

import prep_xlsx_to_sqlite as prep


def _menu_link(ws, coord: str, target_sheet: str, *, label: str | None = None) -> None:
    cell = ws[coord]
    cell.value = label or target_sheet
    cell.hyperlink = Hyperlink(ref=coord, location=f"'{target_sheet}'!A1")


def _main_page_link(ws, coord: str, parent_sheet: str) -> None:
    cell = ws[coord]
    cell.value = "Main Page"
    cell.hyperlink = Hyperlink(ref=coord, location=f"'{parent_sheet}'!A1")


def _add_simple_content_sheet(wb: Workbook, name: str, parent_sheet: str = "Logs Menu") -> None:
    ws = wb.create_sheet(name)
    ws["B1"] = "Name"
    _main_page_link(ws, "F1", parent_sheet)
    ws["A2"] = "SECTION"
    ws["A3"] = "N"
    ws["B3"] = f"{name} Row"


def _build_logs_menu_workbook(path: Path) -> None:
    wb = Workbook()

    logs_menu = wb.active
    logs_menu.title = "Logs Menu"
    logs_menu["A2"] = "Logs Menu"
    _menu_link(logs_menu, "A4", "Logs Menu - Crafting Log")
    _menu_link(logs_menu, "A5", "Logs Menu - Gathering Log")
    _menu_link(logs_menu, "A6", "Hunting Logs")
    _menu_link(logs_menu, "A7", "Sightseeing Logs")
    _menu_link(logs_menu, "A8", "Fishing Logs")
    _menu_link(logs_menu, "A9", "Fish Guide")
    _menu_link(logs_menu, "A10", "Orchestrion")

    crafting = wb.create_sheet("Logs Menu - Crafting Log")
    crafting["A2"] = "Crafting Log"
    _menu_link(crafting, "A4", "Hunting Logs")
    _menu_link(crafting, "A5", "Sightseeing Logs")
    _menu_link(crafting, "A6", "Fishing Logs")
    _menu_link(crafting, "A7", "Fish Guide")
    _menu_link(crafting, "A8", "Orchestrion")
    _menu_link(crafting, "A9", "Master Crafting Books")

    gathering = wb.create_sheet("Logs Menu - Gathering Log")
    gathering["A2"] = "Gathering Log"
    _menu_link(gathering, "A4", "Hunting Logs")
    _menu_link(gathering, "A5", "Sightseeing Logs")
    _menu_link(gathering, "A6", "Fishing Logs")
    _menu_link(gathering, "A7", "Fish Guide")
    _menu_link(gathering, "A8", "Orchestrion")
    _menu_link(gathering, "A9", "Miner Logs")

    for sheet_name in (
        "Hunting Logs",
        "Sightseeing Logs",
        "Fishing Logs",
        "Fish Guide",
        "Orchestrion",
        "Master Crafting Books",
        "Miner Logs",
    ):
        _add_simple_content_sheet(wb, sheet_name)

    wb.save(path)
    wb.close()


def _sheet_parent(conn: sqlite3.Connection, run_id: int, sheet_name: str) -> str | None:
    row = conn.execute(
        "SELECT parent_sheet FROM sheets WHERE run_id = ? AND sheet_name = ?",
        (run_id, sheet_name),
    ).fetchone()
    assert row is not None
    return row["parent_sheet"]


def test_logs_shared_submenu_links_stay_top_level_and_fish_guide_nests(tmp_path: Path):
    xlsx_path = tmp_path / "logs_links.xlsx"
    db_path = tmp_path / "tracker.sqlite"
    _build_logs_menu_workbook(xlsx_path)

    prep.ingest(xlsx_path, db_path)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        run_id = int(conn.execute("SELECT MAX(id) FROM ingest_runs").fetchone()[0])

        # Targets linked by multiple Logs submenus should stay under Logs Menu.
        assert _sheet_parent(conn, run_id, "Hunting Logs") == "Logs Menu"
        assert _sheet_parent(conn, run_id, "Sightseeing Logs") == "Logs Menu"
        assert _sheet_parent(conn, run_id, "Fishing Logs") == "Logs Menu"
        assert _sheet_parent(conn, run_id, "Orchestrion") == "Logs Menu"

        # Submenu-unique targets stay nested under that submenu.
        assert _sheet_parent(conn, run_id, "Master Crafting Books") == "Logs Menu - Crafting Log"
        assert _sheet_parent(conn, run_id, "Miner Logs") == "Logs Menu - Gathering Log"

        # Explicit structural override: Fish Guide sits under Fishing Logs.
        assert _sheet_parent(conn, run_id, "Fish Guide") == "Fishing Logs"
    finally:
        conn.close()
