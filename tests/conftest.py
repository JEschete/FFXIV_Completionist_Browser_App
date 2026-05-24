"""Shared pytest fixtures for the FFXIV tracker test suite.

The whole suite runs against a tiny *synthetic* workbook built in-code with
openpyxl, ingested into a throwaway SQLite DB. Nothing here touches the real
``data/ffxiv_tracker.sqlite`` or the user's progress sidecars — every module
path that points at ``data/`` is monkeypatched onto a tmp directory.

Fixture layering:
  * ``synthetic_workbook`` (session) — writes the .xlsx once.
  * ``template_db`` (session)        — ingests that workbook once.
  * ``ingested_db`` (function)       — copies the template, repoints all the
                                       module-level ``data/`` paths at a tmp
                                       dir, and resets process-wide caches.
  * ``conn`` (function)              — an open connection + the run id.
  * ``client`` (function)            — a FastAPI TestClient over the same DB.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

REPO_ROOT = Path(__file__).resolve().parent.parent
# Make `app`, `scripts/…` and `CharacterScraping/…` importable regardless of
# how pytest is launched.
for _p in (REPO_ROOT, REPO_ROOT / "scripts"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import prep_xlsx_to_sqlite as prep  # noqa: E402  (path set up above)


# --- synthetic workbook -----------------------------------------------------

def _main_page_link(ws, coord: str, parent_sheet: str) -> None:
    """Add a 'Main Page' hyperlink in a far column so the ingest reads it as
    the content sheet's parent link (find_parent_link looks for col >= 6)."""
    cell = ws[coord]
    cell.value = "Main Page"
    cell.hyperlink = Hyperlink(ref=coord, location=f"'{parent_sheet}'!A1")


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    # Workbook() seeds one sheet; repurpose it as the top menu.
    menu = wb.active
    menu.title = "Character Menu"
    menu["A2"] = "Character Menu"
    # Menu hyperlinks to its children (used by the repair / section passes).
    for coord, target in (("A4", "Story Quests"), ("A5", "Side Stuff"), ("A6", "Classes-Jobs")):
        menu[coord] = target
        menu[coord].hyperlink = Hyperlink(ref=coord, location=f"'{target}'!A1")

    # --- a chain content sheet (section label contains "chain") ---
    story = wb.create_sheet("Story Quests")
    story["B1"] = "Quest"
    story["C1"] = "Notes"
    _main_page_link(story, "F1", "Character Menu")
    story["A2"] = "MAIN STORY CHAIN"          # uppercase banner -> section
    story["A3"], story["B3"] = "Y", "Quest Alpha"   # done baseline
    story["A4"], story["B4"] = "N", "Quest Beta"    # todo
    story["A5"], story["B5"] = "N", "Quest Gamma"   # todo

    # --- a plain (non-chain) content sheet ---
    side = wb.create_sheet("Side Stuff")
    side["B1"] = "Name"
    _main_page_link(side, "F1", "Character Menu")
    side["A2"] = "ODDS AND ENDS"
    side["A3"], side["B3"] = "Y", "Thing One"   # done
    side["A4"], side["B4"] = "X", "Thing Two"   # excluded
    side["A5"], side["B5"] = "N", "Thing Three"  # todo

    # --- a value sheet (name triggers VALUE_SHEET_PATTERNS) ---
    jobs = wb.create_sheet("Classes-Jobs")
    jobs["B1"] = "Job"
    jobs["C1"] = "Current Level"
    _main_page_link(jobs, "F1", "Character Menu")
    jobs["A2"] = "TANKS"
    jobs["B3"], jobs["C3"] = "Paladin", 90
    jobs["B4"], jobs["C4"] = "Warrior", 100

    wb.save(path)
    wb.close()


@pytest.fixture(scope="session")
def synthetic_workbook(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("workbook") / "synthetic.xlsx"
    _build_workbook(path)
    return path


@pytest.fixture(scope="session")
def template_db(synthetic_workbook, tmp_path_factory) -> Path:
    """Ingest the synthetic workbook once into a reusable template DB file."""
    db_path = tmp_path_factory.mktemp("template_db") / "template.sqlite"
    prep.ingest(synthetic_workbook, db_path)
    return db_path


# --- per-test isolated DB + repointed data paths ----------------------------

def _reset_module_caches() -> None:
    from app import db, progress_io

    db._VALUE_CAPS_CACHE_MTIME_NS = None
    db._VALUE_CAPS_CACHE_DATA = {}
    progress_io.invalidate_cache()
    with progress_io._sidecar_resolution_lock:
        progress_io._resolved_sidecar_paths.clear()


@pytest.fixture
def ingested_db(template_db, tmp_path, monkeypatch):
    """A fresh copy of the template DB plus tmp data/progress dirs, with every
    module-level ``data/`` path repointed so tests never touch real files."""
    from app import db, progress_io

    data_dir = tmp_path / "data"
    progress_dir = data_dir / "progress"
    progress_dir.mkdir(parents=True)

    db_copy = data_dir / "ffxiv_tracker.sqlite"
    shutil.copy2(template_db, db_copy)

    monkeypatch.setattr(db, "DB_PATH", db_copy)
    monkeypatch.setattr(db, "VALUE_CAPS_PATH", data_dir / "value_caps.json")
    monkeypatch.setattr(progress_io, "PROGRESS_DIR", progress_dir)

    # main.py caches "have we reconciled this run yet" process-wide.
    try:
        import app.main as main_mod

        monkeypatch.setattr(main_mod, "LAST_RECONCILED_RUN_TOKEN", None)
        monkeypatch.setattr(main_mod, "SESSION_BASELINE_SNAPSHOT", None)
        monkeypatch.setattr(main_mod, "LAST_BETWEEN_RUN_REPORT_PATH", None)
    except Exception:
        pass

    _reset_module_caches()
    yield data_dir
    _reset_module_caches()


@pytest.fixture
def conn(ingested_db):
    """Open connection + the latest run id for the isolated DB."""
    from app import db

    connection = db.get_connection()
    run_id = db.latest_run_id(connection)
    assert run_id is not None
    try:
        yield connection, run_id
    finally:
        connection.close()


@pytest.fixture
def character_id(conn):
    """The default 'Adventurer' character created during ingest."""
    connection, _ = conn
    from app import db

    chars = db.fetch_characters(connection)
    assert chars, "ingest should have created a default character"
    return int(chars[0]["id"])


@pytest.fixture
def client(ingested_db):
    """FastAPI TestClient bound to the isolated DB. Entering the context
    manager runs the lifespan (sidecar reconcile) against the tmp DB."""
    from fastapi.testclient import TestClient

    import app.main as main_mod

    with TestClient(main_mod.app) as test_client:
        yield test_client
