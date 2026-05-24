"""Coverage for scripts/prep_xlsx_to_sqlite.py pure helpers + FormulaEvaluator."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

import prep_xlsx_to_sqlite as prep


# --- small pure helpers -----------------------------------------------------

def test_norm_key():
    assert prep.norm_key("Current Level", 0) == "current_level"
    assert prep.norm_key(None, 3) == "col_3"
    assert prep.norm_key("!!!", 4) == "col_4"


def test_norm_value():
    import datetime as dt
    assert prep.norm_value(None) is None
    assert prep.norm_value("  x ") == "x"
    assert prep.norm_value("") is None
    assert prep.norm_value(dt.date(2024, 1, 2)) == "2024-01-02"


def test_parse_state():
    assert prep.parse_state("Y") == "done"
    assert prep.parse_state("X") == "excluded"
    assert prep.parse_state("n/a") == "excluded"
    assert prep.parse_state(None) == "todo"
    assert prep.parse_state("anything") == "todo"


def test_is_inline_section_marker():
    assert prep.is_inline_section_marker("91-100", {}) is True
    assert prep.is_inline_section_marker("(See Shared Craft Log)", {}) is True
    assert prep.is_inline_section_marker("91-100", {"x": "y"}) is False  # has data
    assert prep.is_inline_section_marker(None, {}) is False
    assert prep.is_inline_section_marker("Normal Label", {}) is False


def test_split_candidates():
    assert prep.split_candidates(None) == []
    assert prep.split_candidates("A, B / C and D") == ["A", "B", "C", "D"]
    assert prep.split_candidates("x, y") == []  # filler tokens dropped


def test_menu_helpers():
    assert prep.is_menu_sheet("Character Menu") is True
    assert prep.is_menu_sheet("Duty Menu - Journal") is True
    assert prep.is_menu_sheet("Story Quests") is False
    assert prep.menu_parent("Duty Menu - Journal") == "Duty Menu"
    assert prep.menu_parent("Story Quests") is None


def test_section_is_chain():
    assert prep.section_is_chain("Endwalker Quests", None) is True   # ALWAYS_CHAIN
    assert prep.section_is_chain("Side Stuff", "Some Chain Block") is True  # "chain" in label
    assert prep.section_is_chain("Side Stuff", "Regular") is False


def test_is_numeric_and_row_hash():
    assert prep._is_numeric(5) is True
    assert prep._is_numeric(True) is False
    assert prep._is_numeric("5") is False
    # hash is stable + 12 hex chars
    h = prep._row_hash({"b": 2, "a": 1})
    assert len(h) == 12 and h == prep._row_hash({"a": 1, "b": 2})


def test_pick_columns():
    cols = [
        {"index": 2, "key": "quest", "label": "Quest"},
        {"index": 3, "key": "current_level", "label": "Current Level"},
        {"index": 4, "key": "unlock", "label": "Unlock"},
    ]
    assert prep.pick_value_column(cols)["key"] == "current_level"
    assert prep.pick_label_column(cols, skip=[cols[1]])["key"] == "quest"
    assert prep.pick_unlock_column(cols)["key"] == "unlock"
    assert prep.pick_value_column([{"index": 2, "key": "name", "label": "Name"}]) is None


def test_detect_data_columns_and_link_target():
    wb = Workbook()
    ws = wb.active
    ws["B1"] = "Quest"
    ws["C1"] = "Notes"
    ws["F1"] = "Main Page"
    ws["F1"].hyperlink = Hyperlink(ref="F1", location="'Character Menu'!A1")
    cols = prep.detect_data_columns(ws)
    keys = {c["key"] for c in cols}
    assert "quest" in keys and "notes" in keys  # stops at the hyperlink column
    assert prep.link_target_sheet(ws["F1"]) == "Character Menu"
    assert prep.link_target_sheet(ws["B1"]) is None  # no hyperlink
    wb.close()


def test_cell_fill_and_parent_link():
    wb = Workbook()
    ws = wb.active
    # find_parent_link reads a Main Page hyperlink in the first rows
    ws["F1"] = "Main Page"
    ws["F1"].hyperlink = Hyperlink(ref="F1", location="'Character Menu'!A1")
    assert prep.find_parent_link(ws) == "Character Menu"
    # cell_fill returns "" for an unfilled cell
    assert prep.cell_fill(ws["A1"]) == ""
    wb.close()


def test_process_memory_snapshot_and_checkpoint(capsys):
    snap = prep._process_memory_snapshot_mb()
    assert set(snap.keys()) == {"rss", "private", "peak"}
    prep.log_memory_checkpoint(False, "noop")   # disabled -> no output
    assert capsys.readouterr().out == ""
    prep.log_memory_checkpoint(True, "stage-x")
    assert "stage-x" in capsys.readouterr().out


# --- FormulaEvaluator -------------------------------------------------------

def _evaluator(starting_class: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = 5
    ws["A2"] = "=A1"
    return prep.FormulaEvaluator(wb, starting_class), wb


def test_formula_evaluator_basics():
    ev, wb = _evaluator("GLADIATOR")
    try:
        assert ev.cell_value("S", "A1") == 5
        assert ev.cell_value("S", "A2") == 5            # follows the =A1 ref
        assert ev.cell_value("S", "A1") == 5            # cache hit second time
        assert ev.cell_value("Character Menu", "L2") == "GLADIATOR"  # special ref
        assert ev.cell_value("Missing Sheet", "A1") is None
    finally:
        wb.close()


def test_formula_evaluator_expressions():
    ev, wb = _evaluator("GLADIATOR")
    try:
        assert ev.eval_formula('=IF(\'Character Menu\'!L2="GLADIATOR","Y","N")', "S") == "Y"
        assert ev.eval_formula("=TRUE", "S") is True
        assert ev.eval_formula("=FALSE", "S") is False
        assert ev.eval_formula('="hello"', "S") == "hello"
        assert ev.eval_formula("=3.5", "S") == 3.5
        assert ev.eval_formula("=", "S") is None
        assert ev.eval_formula("=IF(1,2)", "S") is None  # wrong arg count
    finally:
        wb.close()

    ev2, wb2 = _evaluator("MARAUDER")
    try:
        assert ev2.eval_formula('=IF(\'Character Menu\'!L2="GLADIATOR","Y","N")', "S") == "N"
    finally:
        wb2.close()


def test_flatten_refs():
    wb = Workbook()
    ws = wb.active
    ws.title = "Character Menu"
    ws["A1"] = "=L2"
    ws["L2"] = "GLADIATOR"
    blob = prep._flatten_refs(wb, "Character Menu", "A1", depth=3)
    assert "L2" in blob
    assert prep._flatten_refs(wb, "Character Menu", "B5", depth=3) == ""  # not a formula
    wb.close()


def test_collect_class_overrides(tmp_path):
    """A class-conditional A-column formula should emit per-class overrides."""
    wb = Workbook()
    menu = wb.active
    menu.title = "Character Menu"
    menu["L2"] = "GLADIATOR"

    sheet = wb.create_sheet("Test Quests")
    sheet["B1"] = "Quest"
    sheet["A2"] = '=IF(\'Character Menu\'!L2="GLADIATOR","Y","N")'
    sheet["B2"] = "Conditional Quest"
    path = tmp_path / "formula.xlsx"
    wb.save(path)
    wb.close()

    overrides = prep.collect_class_overrides(path, run_id=1)
    # GLADIATOR resolves to done ("Y"); the other 7 classes resolve to todo ("N")
    # -> results differ -> a row is emitted for every class.
    by_class = {cls: state for (_run, cls, _sheet, _row, state) in overrides}
    assert by_class["GLADIATOR"] == "done"
    assert by_class["MARAUDER"] == "todo"
    assert len(overrides) == len(prep.STARTING_CLASSES)
