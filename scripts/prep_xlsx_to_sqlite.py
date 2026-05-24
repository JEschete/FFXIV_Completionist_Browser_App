#!/usr/bin/env python3
"""Ingest the FFXIV completion workbook into SQLite for the web app.

Workbook-driven rules:
- Content-sheet parent is read from the "Main Page" hyperlink.
- Purple banner rows mark sections.
- Column A marker Y/N/X maps to done/todo/excluded.
- Unlock/require columns become explicit edges where possible.
"""

from __future__ import annotations

import argparse
import datetime as dt
import gc
import hashlib
import json
import os
import re
import sqlite3
import subprocess
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[1]))

from app import section_sort


def _row_hash(row_dict: dict) -> str:
    """12-hex-char SHA-256 of a row dict's normalized JSON. Mirrors
    progress_io._hash_row so the per-row identity matches whether the hash
    was generated at ingest time or in the running app."""
    norm = json.dumps(row_dict, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(norm.encode("utf-8")).hexdigest()[:12]

# --- workbook conventions ---------------------------------------------------

SECTION_FILL = "CC66FF"  # purple banner rows

TOP_MENUS = (
    "Character Menu",
    "Duty Menu",
    "Logs Menu",
    "Travel Menu",
    "Social Menu",
)

# submenu name prefix -> parent top menu
MENU_PREFIX_PARENT = {
    "Char. Menu - ": "Character Menu",
    "Duty Menu - ": "Duty Menu",
    "Logs Menu - ": "Logs Menu",
}

# Explicit structural overrides for content sheets whose workbook "Main Page"
# links don't reflect the desired sidebar hierarchy.
CONTENT_PARENT_OVERRIDES: dict[str, str] = {
    "Fish Guide": "Fishing Logs",
}

# data-column header keywords used to pick the "name" of a row
LABEL_KEYS = (
    "job",
    "class",
    "quest",
    "name",
    "title",
    "item",
    "achievement",
    "spell",
    "mount",
    "minion",
    "emote",
    "card",
    "orchestrion",
    "weapon",
    "tool",
    "barding",
    "fish",
    "log",
    "action",
    "voice",
    "currency",
    "building",
    "rank",
)

# headers that identify the numeric column on a per-row fillable sheet
VALUE_KEYS = ("current_level",)

# data-column header keywords that describe what a row unlocks / requires
UNLOCK_KEYS = ("unlock", "requires", "required", "prereq", "next")

# Sheets where each row carries its own user-entered numeric value (e.g. each
# job's current level). NOT rank ladders -- those are checkbox per tier.
VALUE_SHEET_PATTERNS = ("classes-jobs",)

# Sheets whose every section is treated as a sequential prerequisite chain
# (MSQ + raid alliance stories + relic paths + per-class job-quest tracks).
# Anything outside this set falls back to the "section name contains 'Chain'"
# heuristic — so sidequest/FATE/log sheets emit *no* sequence edges and don't
# trigger cascading completion.
ALWAYS_CHAIN_SHEETS = frozenset({
    # Main Scenario, expansion by expansion
    "Seventh Umbral Era Quests", "Seventh Astral Era Quests",
    "Heavensward Quests", "Dragonsong Quests", "Post-Dragonsong Quests",
    "Stormblood Quests", "Post-Stormblood Quests",
    "Shadowbringers Quests", "Post-Shadowbringers Quests",
    "Post-Shadowbringers Quests II",
    "Endwalker Quests", "Post-Endwalker Quests",
    "Dawntrail Quests", "Post-Dawntrail Quests", "Post-Dawntrail Quests II",
    # Chronicles of a New Era — each raid storyline is linear
    "Primals", "Bahamut", "The Crystal Tower", "Alexander", "The Warring Triad",
    "The Shadow of Mach", "Omega", "Return to Ivalice", "The Four Lords",
    "Eden", "YoRHa Dark Apocalypse", "The Sorrow of Werlyt",
    "Pandæmonium", "Myths of the Realm", "The Arcadion",
    "Echoes of Vanadiel",
    # Linear side stories
    "Chronicles of Light", "Hildibrand", "Weapon Enhancement",
    "Records of Unusual Endeavors", "Side Story Quests",
    # Class / job / role quests — each section (per class) is a linear path
    "Disciple of War Quests", "Disciple of Magic Quests",
    "Disciple of the Hand Quests", "Disciple of the Land Quests",
    "Disciple of War Job Quests", "Disciple of Magic Job Quests",
    "Role Quests", "Hall of the Novice",
    "Crystalline Mean Quests", "Studium Quests", "Wachumeqimeqi Quests",
    # Relic upgrade paths — each weapon goes through stages sequentially
    "Relic Tools", "Relic Weapons",
})


def section_is_chain(sheet_name: str, section_label: str | None) -> bool:
    """A row is part of a real prerequisite chain when its sheet is end-to-end
    sequential, or when the workbook explicitly named the section a Chain."""
    if sheet_name in ALWAYS_CHAIN_SHEETS:
        return True
    if section_label and "chain" in section_label.lower():
        return True
    return False


# Within a chain section, restart the prerequisite chain whenever this column's
# value changes between consecutive rows. Lets the workbook keep one broad
# banner ("A Realm Reborn Tools (Lucis)", "Disciple of War") covering several
# *independent* sub-chains — per crafter, per class, per faculty — without
# inventing fake sub-banners. Without this, the ingest would chain every
# Carpenter tool into every Blacksmith tool inside a shared section.
SUB_CHAIN_BOUNDARY_COLUMN: dict[str, str] = {
    "Relic Tools": "job",
    "Relic Weapons": "job",
    "Disciple of War Quests": "class",
    "Disciple of Magic Quests": "class",
    "Disciple of the Hand Quests": "class",
    "Disciple of the Land Quests": "class",
    "Disciple of War Job Quests": "job",
    "Disciple of Magic Job Quests": "job",
    "Studium Quests": "faculty",
    "Hall of the Novice": "class",
}

# The 8 ARR starting classes the workbook's class-conditional formulas key on.
STARTING_CLASSES = (
    "ARCANIST",
    "ARCHER",
    "CONJURER",
    "GLADIATOR",
    "LANCER",
    "MARAUDER",
    "PUGILIST",
    "THAUMATURGE",
)

# The cell the workbook reads for the active starting class.
STARTING_CLASS_REF = ("Character Menu", "L2")

# Sheets that are reference/info only (not trackable content rows).
READONLY_SHEETS = {
    "Read Me",
    "Before You Get Started, Take A Moment To Look Over The Following Information.",
}

# Workbook tabs to exclude from DB import entirely.
# These are informational pages, not navigable tracker content.
SKIP_IMPORT_SHEETS = {
    "Read Me",
}


# --- helpers ----------------------------------------------------------------

def norm_key(value, fallback_index: int) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or f"col_{fallback_index}"


def norm_value(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def cell_fill(cell) -> str:
    try:
        rgb = cell.fill.start_color.rgb
        if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
            return rgb[-6:].upper()
    except Exception:
        pass
    return ""


def parse_state(raw: str | None) -> str:
    v = (raw or "").strip().lower()
    if v in ("y", "yes", "x_done", "done", "complete", "completed", "true"):
        return "done"
    if v in ("x", "excluded", "n/a", "na"):
        return "excluded"
    return "todo"


_INLINE_SECTION_RANGE_RE = re.compile(r"^\d+\s*-\s*\d+$")
_INLINE_SECTION_NOTE_RE = re.compile(r"^\(\s*see\s+shared\s+craft\s+log\s*\)$", re.IGNORECASE)


def is_inline_section_marker(a_val: str | None, data: dict[str, str]) -> bool:
    """Rows with only a range/note marker in column A should be section headers.

    Example markers seen in crafting logs: "91-100", "81-90",
    "(See Shared Craft Log)".
    """
    if a_val is None or data:
        return False
    text = a_val.strip()
    if not text:
        return False
    return bool(
        _INLINE_SECTION_RANGE_RE.fullmatch(text)
        or _INLINE_SECTION_NOTE_RE.fullmatch(text)
    )


def link_target_sheet(cell) -> str | None:
    """Return the sheet name a hyperlink points at, or None."""
    hl = cell.hyperlink
    if hl is None:
        return None
    loc = hl.location or hl.target or ""
    m = re.match(r"^'?(.*?)'?!", str(loc))
    target = m.group(1) if m else str(loc)
    if not target or target.lower() == "null":
        return None
    return target


def split_candidates(value: str | None) -> list[str]:
    if not value:
        return []
    parts = re.split(r"\s*(?:,|/|;|\||->|=>|\band\b)\s*", value, flags=re.IGNORECASE)
    out: list[str] = []
    for p in parts:
        p = p.strip()
        if p and p.lower() not in {"x", "n", "y", "yes", "no", "-"}:
            out.append(p)
    return out


_TASKLIST_CACHE_TS: float = 0.0
_TASKLIST_CACHE_MB: float | None = None


def _tasklist_working_set_mb() -> float | None:
    """Fallback RSS from tasklist (Windows only), cached briefly.

    Spawning tasklist on every checkpoint would be expensive, so cache for a
    short interval to keep mem-log overhead low.
    """
    global _TASKLIST_CACHE_TS, _TASKLIST_CACHE_MB
    now = time.monotonic()
    if now - _TASKLIST_CACHE_TS < 0.7:
        return _TASKLIST_CACHE_MB

    _TASKLIST_CACHE_TS = now
    _TASKLIST_CACHE_MB = None
    if not sys.platform.startswith("win"):
        return None

    try:
        proc = subprocess.run(
            [
                "tasklist",
                "/FI",
                f"PID eq {os.getpid()}",
                "/FO",
                "CSV",
                "/NH",
            ],
            capture_output=True,
            text=True,
            timeout=3,
            check=False,
        )
        line = (proc.stdout or "").strip().splitlines()
        if not line:
            return None
        row = line[0]
        # CSV looks like: "python.exe","1234","Console","1","21,384 K"
        parts = [p.strip().strip('"') for p in row.split(",")]
        if len(parts) < 5:
            return None
        mem_raw = parts[-1].upper().replace(" K", "").replace(",", "").strip()
        kb = float(mem_raw)
        _TASKLIST_CACHE_MB = kb / 1024.0
    except Exception:
        _TASKLIST_CACHE_MB = None
    return _TASKLIST_CACHE_MB


def _process_memory_snapshot_mb() -> dict[str, float | None]:
    """Best-effort current-process memory snapshot in MB.

    On Windows this reports current working set + private bytes + peak working
    set. On other platforms, only peak RSS may be available.
    """
    out: dict[str, float | None] = {
        "rss": None,
        "private": None,
        "peak": None,
    }

    if sys.platform.startswith("win"):
        try:
            import ctypes
            from ctypes import wintypes

            class PROCESS_MEMORY_COUNTERS_EX(ctypes.Structure):
                _fields_ = [
                    ("cb", wintypes.DWORD),
                    ("PageFaultCount", wintypes.DWORD),
                    ("PeakWorkingSetSize", ctypes.c_size_t),
                    ("WorkingSetSize", ctypes.c_size_t),
                    ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                    ("PagefileUsage", ctypes.c_size_t),
                    ("PeakPagefileUsage", ctypes.c_size_t),
                    ("PrivateUsage", ctypes.c_size_t),
                ]

            psapi = ctypes.WinDLL("psapi")
            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)

            # Explicit signatures make the call reliable across Python builds.
            kernel32.GetCurrentProcess.restype = wintypes.HANDLE
            psapi.GetProcessMemoryInfo.argtypes = [
                wintypes.HANDLE,
                ctypes.c_void_p,
                wintypes.DWORD,
            ]
            psapi.GetProcessMemoryInfo.restype = wintypes.BOOL

            h_process = kernel32.GetCurrentProcess()
            counters = PROCESS_MEMORY_COUNTERS_EX()
            counters.cb = ctypes.sizeof(counters)
            ok = psapi.GetProcessMemoryInfo(
                h_process,
                ctypes.byref(counters),
                counters.cb,
            )
            if ok:
                mb = 1024.0 * 1024.0
                out["rss"] = float(counters.WorkingSetSize) / mb
                out["private"] = float(counters.PrivateUsage) / mb
                out["peak"] = float(counters.PeakWorkingSetSize) / mb
                return out
        except Exception:
            pass

        # Last-resort Windows fallback: tasklist working set only.
        rss_mb = _tasklist_working_set_mb()
        if rss_mb is not None:
            out["rss"] = rss_mb
            return out

    # Non-Windows fallback: peak RSS only (platform dependent units).
    try:
        import resource

        getrusage = getattr(resource, "getrusage", None)
        rusage_self = getattr(resource, "RUSAGE_SELF", None)
        if callable(getrusage) and rusage_self is not None:
            usage = getrusage(rusage_self)
            rss = float(getattr(usage, "ru_maxrss", 0.0))
            # macOS reports bytes; Linux reports KB.
            if sys.platform == "darwin":
                out["peak"] = rss / (1024.0 * 1024.0)
            else:
                out["peak"] = rss / 1024.0
    except Exception:
        pass
    return out


def log_memory_checkpoint(enabled: bool, stage: str) -> None:
    if not enabled:
        return
    snap = _process_memory_snapshot_mb()
    ts = dt.datetime.now().strftime("%H:%M:%S")

    def _fmt(value: float | None) -> str:
        return "n/a" if value is None else f"{value:,.1f}MB"

    print(
        f"[mem {ts}] {stage} | "
        f"rss={_fmt(snap.get('rss'))} "
        f"private={_fmt(snap.get('private'))} "
        f"peak={_fmt(snap.get('peak'))}"
    )


# --- structural analysis ----------------------------------------------------

def is_menu_sheet(name: str) -> bool:
    return name in TOP_MENUS or any(name.startswith(p) for p in MENU_PREFIX_PARENT)


def menu_parent(name: str) -> str | None:
    for prefix, parent in MENU_PREFIX_PARENT.items():
        if name.startswith(prefix):
            return parent
    return None


def find_parent_link(ws) -> str | None:
    """Content sheets carry a 'Main Page' hyperlink back to their parent menu."""
    max_col = min(ws.max_column or 0, 80)
    if max_col < 1:
        return None
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=max_col):
        for cell in row:
            if cell.hyperlink is None:
                continue
            label = str(cell.value or "").lower()
            target = link_target_sheet(cell)
            if target and ("main" in label or cell.column >= 6):
                return target
    return None


def _is_numeric(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _has_percent_left(ws, row: int, col: int, max_gap: int = 2) -> bool:
    """True if there's a numeric cell within ``max_gap`` columns to the left
    of (row, col). The workbook leaves a one-column gap between the percent
    and its text label (e.g. G=0  H=blank  I='Main Scenario'), so checking
    only the immediate left column misses every pair."""
    for offset in range(1, max_gap + 1):
        c = col - offset
        if c < 1:
            return False
        v = ws.cell(row=row, column=c).value
        if _is_numeric(v):
            return True
        # if we hit another *text* cell on the way left, the percent (if any)
        # belongs to a different group, not this one
        if isinstance(v, str) and v.strip():
            return False
    return False


def extract_menu_groups(
    ws, is_real_sheet=None
) -> dict[str, str]:
    """Walk a menu sheet's columnar layout and return a mapping of
    ``child_label_lower -> group_label`` for every text-cell that sits in a
    "child entry" position (i.e. has a numeric % cell within ~2 columns to
    its left, just like the column header above it).

    The workbook organizes most menu sheets into several side-by-side
    columns (e.g. Duty Menu - Journal: Main Scenario / Sidequests / Allied
    Society Quests / Other Quests). Each column has a header row of text
    cells whose left neighbor is a numeric "% complete". Below that,
    children of the same group occupy the same column with the same
    layout. We use the consistent percent-on-the-left signal to:

      1. Detect the header row (>= 2 such pairs).
      2. Pick out the (text-column, header) pairs.
      3. Walk subsequent rows, grouping each child under its column's header.

    If ``is_real_sheet`` is provided, any entry that looks like a child by
    layout (has a percent cell on the left) but doesn't resolve to a real
    sheet name is treated as an *in-column sub-header* — it doesn't itself
    get a section assignment, but it updates the running group label for
    subsequent children in that column. This is how Duty Menu - Journal's
    "Chronicles of a New Era", "Class & Job Quests", and "Levequests" rows
    get promoted to top-level sections instead of being mistaken for child
    sheets.

    Returns ``{}`` for menu sheets that don't have a multi-column layout
    (callers leave parent_menu_section NULL — UI renders ungrouped)."""
    if ws.max_row is None or ws.max_row < 8:
        return {}
    max_scan_col = min(ws.max_column or 0, 80)
    if max_scan_col < 1:
        return {}

    # 1) find the header row: scan rows 5-20, pick the row with the most
    # "text cell with a percent cell within ~2 columns to its left" pairs
    # (>= 2 pairs). The 2-column tolerance is because the workbook leaves
    # an empty column between each (percent, text) pair (G/H/I, L/M/N, …).
    header_row_idx: int | None = None
    header_pairs: list[tuple[int, str]] = []
    for r_idx in range(5, min(ws.max_row, 25) + 1):
        pairs: list[tuple[int, str]] = []
        for col_idx in range(1, max_scan_col + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            if not text or col_idx < 2:
                continue
            if _has_percent_left(ws, r_idx, col_idx):
                pairs.append((col_idx, text))
        if len(pairs) >= 2 and len(pairs) > len(header_pairs):
            header_row_idx = r_idx
            header_pairs = pairs

    if header_row_idx is None or not header_pairs:
        return {}

    # 2) walk down each column collecting child entries. A child entry is
    # any text cell in a header column with the same percent-on-the-left
    # signal as the header itself. All-caps short strings *without* that
    # signal are treated as in-column sub-headers and update the running
    # label for that column (e.g. X16='LEVEQUESTS' inside Other Quests).
    columns: dict[int, str] = dict(header_pairs)
    out: dict[str, str] = {}

    for r_idx in range(header_row_idx + 1, (ws.max_row or 0) + 1):
        for col_idx, current_label in list(columns.items()):
            cell = ws.cell(row=r_idx, column=col_idx)
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            if not text:
                continue
            has_pct = _has_percent_left(ws, r_idx, col_idx)
            if has_pct and (is_real_sheet is None or is_real_sheet(text)):
                # genuine child entry — record under the column's current label
                out.setdefault(text.lower(), current_label)
            elif has_pct and is_real_sheet is not None:
                # has the child-layout shape but doesn't name a real sheet —
                # it's an in-column sub-header (e.g. "Chronicles of a New
                # Era", "Class & Job Quests", "Levequests"). Promote it to
                # the running group label for everything below it.
                columns[col_idx] = text
            elif text.isupper() and 3 < len(text) < 50:
                columns[col_idx] = text.title()

    return out


def extract_menu_link_targets(ws) -> set[str]:
    """Return every sheet name directly referenced by hyperlinks in a menu sheet.

    Used as a structural source of truth for child-parent relationships when a
    content sheet's own "Main Page" link points to a top menu instead of the
    submenu that actually contains it."""
    targets: set[str] = set()
    max_col = min(ws.max_column or 0, 80)
    if max_col < 1:
        return targets
    for row in ws.iter_rows(min_row=1, max_row=(ws.max_row or 0), min_col=1, max_col=max_col):
        for cell in row:
            target = link_target_sheet(cell)
            if target:
                targets.add(target)
    return targets


def detect_data_columns(ws) -> list[dict]:
    """Detect non-empty data headers from B onward until the sidebar link."""
    cols: list[dict] = []
    max_scan = min(max(ws.max_column or 0, 6), 14)
    for idx in range(2, max_scan + 1):
        cell = ws.cell(row=1, column=idx)
        if cell.hyperlink is not None:
            break
        text = norm_value(cell.value)
        if text is None:
            continue
        cols.append(
            {
                "index": idx,
                "letter": get_column_letter(idx),
                "key": norm_key(text, idx),
                "label": text,
            }
        )

    # de-dupe keys
    seen: dict[str, int] = {}
    for c in cols:
        key = c["key"]
        n = seen.get(key, 0)
        if n:
            c["key"] = f"{key}_{n + 1}"
        seen[c["key"]] = n + 1
    return cols


def pick_label_column(columns: list[dict], skip: list[dict] | None = None) -> dict | None:
    skip_keys = {c["key"] for c in (skip or []) if c}
    for key in LABEL_KEYS:
        for c in columns:
            if c["key"] in skip_keys:
                continue
            if key in c["key"]:
                return c
    for c in columns:
        if c["key"] not in skip_keys:
            return c
    return None


def pick_value_column(columns: list[dict]) -> dict | None:
    for key in VALUE_KEYS:
        for c in columns:
            if key in c["key"]:
                return c
    return None


def pick_unlock_column(columns: list[dict]) -> dict | None:
    for c in columns:
        if any(k in c["key"] for k in UNLOCK_KEYS):
            return c
    return None


# --- schema -----------------------------------------------------------------

SCHEMA = """
PRAGMA journal_mode=WAL;

CREATE TABLE ingest_runs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_file TEXT NOT NULL,
    started_at TEXT NOT NULL,
    completed_at TEXT,
    sheet_count INTEGER NOT NULL DEFAULT 0,
    row_count INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE sheets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id INTEGER NOT NULL,
    sheet_index INTEGER NOT NULL,
    sheet_name TEXT NOT NULL,
    title TEXT NOT NULL,
    is_menu INTEGER NOT NULL DEFAULT 0,
    is_readonly INTEGER NOT NULL DEFAULT 0,
    parent_sheet TEXT,
    -- Sub-grouping within a parent menu sheet. Many menu sheets organize
    -- their child links into multiple columns under category headers
    -- (e.g. Duty Menu - Journal: Main Scenario / Sidequests / Allied Society
    -- / Other Quests). NULL means this sheet wasn't matched to any column
    -- header during ingest -- it'll render ungrouped.
    parent_menu_section TEXT,
    data_columns_json TEXT NOT NULL,
    label_key TEXT,
    value_key TEXT,
    total_rows INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE nodes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id INTEGER NOT NULL,
    sheet_name TEXT NOT NULL,
    row_index INTEGER NOT NULL,
    label TEXT,
    baseline_state TEXT NOT NULL DEFAULT 'todo',
    row_type TEXT NOT NULL DEFAULT 'checkbox',
    section_label TEXT,
    seq INTEGER NOT NULL DEFAULT 0,
    row_json TEXT NOT NULL,
    -- 12-hex-char SHA-256 of the row's normalized JSON; used as the
    -- "content fingerprint" tier of progress reconciliation when label
    -- and section have both shifted.
    stable_hash TEXT
);

CREATE TABLE edges (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id INTEGER NOT NULL,
    sheet_name TEXT NOT NULL,
    edge_type TEXT NOT NULL,
    source_row_index INTEGER,
    source_label TEXT,
    target_row_index INTEGER,
    target_label TEXT,
    resolved INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE characters (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    starting_class TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE class_overrides (
    run_id         INTEGER NOT NULL,
    starting_class TEXT NOT NULL,
    sheet_name     TEXT NOT NULL,
    row_index      INTEGER NOT NULL,
    state          TEXT NOT NULL CHECK(state IN ('done','todo','excluded')),
    PRIMARY KEY (run_id, starting_class, sheet_name, row_index)
);

CREATE INDEX idx_class_overrides_lookup
    ON class_overrides(run_id, starting_class, sheet_name);

CREATE TABLE character_progress (
    character_id INTEGER NOT NULL,
    run_id INTEGER NOT NULL,
    sheet_name TEXT NOT NULL,
    row_index INTEGER NOT NULL,
    state TEXT NOT NULL CHECK(state IN ('done','todo','excluded')),
    progress_percent REAL,
    updated_at TEXT NOT NULL,
    PRIMARY KEY (character_id, run_id, sheet_name, row_index)
);

CREATE TABLE progress_rollup (
    character_id INTEGER NOT NULL,
    run_id       INTEGER NOT NULL,
    sheet_name   TEXT NOT NULL,
    done         INTEGER NOT NULL DEFAULT 0,
    excluded     INTEGER NOT NULL DEFAULT 0,
    total        INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY (character_id, run_id, sheet_name)
);

CREATE INDEX idx_nodes_sheet ON nodes(run_id, sheet_name, row_index);
CREATE INDEX idx_edges_sheet ON edges(run_id, sheet_name, edge_type);
CREATE INDEX idx_progress ON character_progress(character_id, run_id, sheet_name);

-- progress_io reconciliation indexes: tier 1 (sheet+section+label),
-- tier 2 (sheet+label), tier 3 (sheet+content-hash). Tier 4 (row_index)
-- is already covered by idx_nodes_sheet.
CREATE INDEX idx_nodes_section_label
    ON nodes(run_id, sheet_name, section_label, label);
CREATE INDEX idx_nodes_label
    ON nodes(run_id, sheet_name, label);
CREATE INDEX idx_nodes_stable_hash
    ON nodes(run_id, sheet_name, stable_hash);
"""


# --- starting-class formula evaluation --------------------------------------

class FormulaEvaluator:
    """Tiny Excel formula interpreter for class-conditional IF/equality/cell refs."""

    def __init__(self, wb, starting_class: str):
        self.wb = wb
        self.starting_class = starting_class
        self._cache: dict[tuple[str, str], object] = {}

    def cell_value(self, sheet_name: str, coord: str):
        key = (sheet_name, coord)
        if key in self._cache:
            return self._cache[key]
        if (sheet_name, coord) == STARTING_CLASS_REF:
            self._cache[key] = self.starting_class
            return self.starting_class
        try:
            cell = self.wb[sheet_name][coord]
        except (KeyError, ValueError):
            return None
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            try:
                v = self.eval_formula(v, sheet_name)
            except Exception:
                v = None
        self._cache[key] = v
        return v

    def eval_formula(self, formula: str, current_sheet: str):
        return self._eval_expr(formula[1:].strip(), current_sheet)

    def _eval_expr(self, expr: str, sheet: str):
        expr = expr.strip()
        if not expr:
            return None

        u = expr.upper()
        if u.startswith("IF(") and expr.endswith(")"):
            args = self._split_args(expr[3:-1])
            if len(args) != 3:
                return None
            cond = self._eval_expr(args[0], sheet)
            return self._eval_expr(args[1] if cond else args[2], sheet)

        lhs, rhs = self._split_compare(expr)
        if lhs is not None and rhs is not None:
            return self._eval_expr(lhs, sheet) == self._eval_expr(rhs, sheet)

        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]
        if u == "TRUE":
            return True
        if u == "FALSE":
            return False

        m = re.match(r"^'([^']+)'!\$?([A-Z]+)\$?(\d+)$", expr)
        if m:
            return self.cell_value(m.group(1), m.group(2) + m.group(3))

        m = re.match(r"^\$?([A-Z]+)\$?(\d+)$", expr)
        if m:
            return self.cell_value(sheet, m.group(1) + m.group(2))

        try:
            return float(expr)
        except ValueError:
            return expr

    @staticmethod
    def _split_args(body: str) -> list[str]:
        depth = 0
        in_quote = False
        cur: list[str] = []
        out: list[str] = []
        for ch in body:
            if ch == '"':
                in_quote = not in_quote
                cur.append(ch)
            elif in_quote:
                cur.append(ch)
            elif ch == "(":
                depth += 1
                cur.append(ch)
            elif ch == ")":
                depth -= 1
                cur.append(ch)
            elif ch == "," and depth == 0:
                out.append("".join(cur).strip())
                cur = []
            else:
                cur.append(ch)
        if cur:
            out.append("".join(cur).strip())
        return out

    @staticmethod
    def _split_compare(expr: str) -> tuple[str | None, str | None]:
        depth = 0
        in_quote = False
        for i, ch in enumerate(expr):
            if ch == '"':
                in_quote = not in_quote
            elif in_quote:
                continue
            elif ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "=" and depth == 0:
                return expr[:i].strip(), expr[i + 1 :].strip()
        return None, None


def _flatten_refs(wb, sheet_name: str, coord: str, depth: int) -> str:
    """Concatenate formula text reachable from a cell within a few hops."""
    if depth <= 0:
        return ""
    try:
        v = wb[sheet_name][coord].value
    except (KeyError, ValueError):
        return ""
    if not isinstance(v, str) or not v.startswith("="):
        return ""

    out = [v]
    for sn, ref in re.findall(r"'([^']+)'!\$?([A-Z]+\d+)", v):
        out.append(_flatten_refs(wb, sn, ref, depth - 1))
    for ref in re.findall(r"(?<!')\b([A-Z]+\d+)\b", v):
        out.append(_flatten_refs(wb, sheet_name, ref, depth - 1))
    return " ".join(out)


def collect_class_overrides(
    xlsx_path: Path, run_id: int, *, mem_log: bool = False
) -> list[tuple[int, str, str, int, str]]:
    """Evaluate class-dependent formula cells and emit per-class state overrides."""
    log_memory_checkpoint(mem_log, "formula-pass: load workbook data_only=False")
    wb = load_workbook(filename=xlsx_path, data_only=False)
    try:
        log_memory_checkpoint(mem_log, "formula-pass: workbook loaded")
        formula_cells: list[tuple[str, int]] = []
        total_sheets = len(wb.sheetnames)
        for idx, name in enumerate(wb.sheetnames, start=1):
            if is_menu_sheet(name) or name in READONLY_SHEETS:
                continue
            ws = wb[name]
            if ws.sheet_state != "visible":
                continue
            for row_idx in range(2, min(ws.max_row or 0, 2000) + 1):
                a = ws.cell(row=row_idx, column=1)
                v = a.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                if "L2" not in _flatten_refs(wb, name, a.coordinate, depth=3):
                    continue
                formula_cells.append((name, row_idx))
            if mem_log and (idx == 1 or idx % 20 == 0 or idx == total_sheets):
                log_memory_checkpoint(
                    mem_log,
                    f"formula-pass: scanned sheet {idx}/{total_sheets} ({name})",
                )

        log_memory_checkpoint(
            mem_log,
            f"formula-pass: candidate formula cells={len(formula_cells)}",
        )

        per_class: dict[str, dict[tuple[str, int], str]] = {}
        for class_idx, cls in enumerate(STARTING_CLASSES, start=1):
            ev = FormulaEvaluator(wb, cls)
            per_class[cls] = {}
            for sheet_name, row_idx in formula_cells:
                coord = f"A{row_idx}"
                val = ev.cell_value(sheet_name, coord)
                if val in ("X", "N", "Y"):
                    state = {"X": "excluded", "N": "todo", "Y": "done"}[val]
                    per_class[cls][(sheet_name, row_idx)] = state
            log_memory_checkpoint(
                mem_log,
                f"formula-pass: evaluated class {class_idx}/{len(STARTING_CLASSES)} ({cls})",
            )

        rows: list[tuple[int, str, str, int, str]] = []
        all_cells = {k for d in per_class.values() for k in d}
        for key in all_cells:
            results = {cls: per_class[cls].get(key) for cls in STARTING_CLASSES}
            if len(set(results.values())) <= 1:
                continue
            sheet_name, row_idx = key
            for cls, state in results.items():
                if state is None:
                    continue
                rows.append((run_id, cls, sheet_name, row_idx, state))
        log_memory_checkpoint(
            mem_log,
            f"formula-pass: overrides built rows={len(rows)}",
        )
        return rows
    finally:
        log_memory_checkpoint(mem_log, "formula-pass: closing workbook")
        try:
            wb.close()
        except Exception:
            pass
        log_memory_checkpoint(mem_log, "formula-pass: workbook closed")


# --- schema rebuild ---------------------------------------------------------

def rebuild_schema(conn: sqlite3.Connection) -> tuple[list[tuple], list[tuple]]:
    """Drop data tables and return saved characters/progress for migration."""
    saved_chars: list[tuple] = []
    saved_progress: list[tuple] = []

    existing = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    if "characters" in existing:
        # Try the newer 4-column form first; fall back if the existing DB
        # predates the starting_class column. The reinsert below pads 3-tuples
        # to 4 by inserting NULL for starting_class.
        try:
            saved_chars = list(conn.execute(
                "SELECT id, name, starting_class, created_at FROM characters"
            ))
        except sqlite3.OperationalError:
            saved_chars = list(conn.execute(
                "SELECT id, name, created_at FROM characters"
            ))
    if "character_progress" in existing:
        try:
            saved_progress = list(
                conn.execute(
                    """
                    SELECT character_id, sheet_name, row_index,
                           state, progress_percent, updated_at
                    FROM (
                        SELECT character_id, sheet_name, row_index,
                               state, progress_percent, updated_at,
                               ROW_NUMBER() OVER (
                                   PARTITION BY character_id, sheet_name, row_index
                                   ORDER BY COALESCE(updated_at, '') DESC, run_id DESC
                               ) AS rn
                        FROM character_progress
                    ) ranked
                    WHERE rn = 1
                    """
                )
            )
        except sqlite3.OperationalError:
            try:
                saved_progress = list(
                    conn.execute(
                        """
                        SELECT character_id, sheet_name, row_index,
                               state, progress_percent, updated_at
                        FROM character_progress
                        WHERE run_id = (SELECT MAX(run_id) FROM character_progress)
                        """
                    )
                )
            except sqlite3.OperationalError:
                saved_progress = []

    for table in (
        "edges",
        "nodes",
        "sheet_rows",
        "sheet_cells",
        "chain_edges",
        "chain_nodes",
        "sheets",
        "ingest_runs",
        "progress_rollup",
        "class_overrides",
        "character_progress",
        "characters",
    ):
        conn.execute(f"DROP TABLE IF EXISTS {table}")

    conn.executescript(SCHEMA)
    conn.commit()
    return saved_chars, saved_progress


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return bool(row)


def _maybe_capture_pre_ingest_baseline(conn: sqlite3.Connection) -> None:
    required_tables = ("characters", "nodes", "character_progress")
    if not all(_table_exists(conn, name) for name in required_tables):
        return
    try:
        from app import progress_report

        snapshot = progress_report.build_snapshot(conn, source="ingest-script-pre-rebuild")
        if snapshot.get("characters"):
            progress_report.save_snapshot(snapshot, progress_report.BASELINE_PATH)
    except Exception as exc:
        print(f"[warn] Skipped pre-ingest baseline snapshot: {exc}")


# --- ingest -----------------------------------------------------------------

def ingest(xlsx_path: Path, db_path: Path, *, mem_log: bool = False) -> None:
    log_memory_checkpoint(mem_log, "ingest: start")
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)

    _maybe_capture_pre_ingest_baseline(conn)

    saved_chars, saved_progress = rebuild_schema(conn)
    log_memory_checkpoint(mem_log, "ingest: schema rebuilt")

    started = dt.datetime.now().isoformat(timespec="seconds")
    raw_run_id = conn.execute(
        "INSERT INTO ingest_runs (source_file, started_at) VALUES (?, ?)",
        (str(xlsx_path), started),
    ).lastrowid
    if raw_run_id is None:
        raise RuntimeError("Failed to insert ingest_runs row")
    run_id = int(raw_run_id)

    print(f"Loading workbook {xlsx_path} ...")
    log_memory_checkpoint(mem_log, "ingest: loading workbook data_only=True")
    wb = load_workbook(filename=xlsx_path, data_only=True)
    log_memory_checkpoint(mem_log, "ingest: workbook loaded")
    sheet_names = set(wb.sheetnames)
    sheet_count = len(wb.sheetnames)

    total_rows = 0
    for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
        if sheet_name in SKIP_IMPORT_SHEETS:
            print(f"  [{sheet_index:3}] {sheet_name:<52} skipped")
            continue

        ws = wb[sheet_name]
        is_menu = is_menu_sheet(sheet_name)
        is_readonly = sheet_name in READONLY_SHEETS
        log_memory_checkpoint(
            mem_log,
            f"sheet {sheet_index}/{sheet_count}: start {sheet_name}",
        )

        if is_menu:
            parent = menu_parent(sheet_name)
        elif is_readonly:
            parent = None
        else:
            parent = find_parent_link(ws)
            if parent and parent not in sheet_names:
                parent = None

        columns = detect_data_columns(ws)
        is_value_sheet = any(p in sheet_name.lower() for p in VALUE_SHEET_PATTERNS)
        value_col = pick_value_column(columns) if is_value_sheet else None
        label_col = pick_label_column(columns, skip=[value_col] if value_col else None)
        unlock_col = pick_unlock_column(columns)
        label_key = label_col["key"] if label_col else None
        value_key = value_col["key"] if value_col else None

        if is_menu or is_readonly:
            title = sheet_name
            banner = norm_value(ws.cell(row=2, column=1).value)
            if banner:
                title = banner.title()
            conn.execute(
                """
                INSERT INTO sheets (
                    run_id, sheet_index, sheet_name, title,
                    is_menu, is_readonly, parent_sheet, data_columns_json,
                    label_key, value_key, total_rows
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    sheet_index,
                    sheet_name,
                    title,
                    int(is_menu),
                    int(is_readonly),
                    parent,
                    json.dumps(columns),
                    label_key,
                    value_key,
                    0,
                ),
            )
            print(f"  [{sheet_index:3}] {sheet_name:<52} menu/readonly")
            log_memory_checkpoint(
                mem_log,
                f"sheet {sheet_index}/{sheet_count}: menu/readonly done {sheet_name}",
            )
            continue

        node_rows: list[tuple] = []
        seq_edges: list[tuple] = []
        unlock_edges: list[tuple] = []
        label_to_row: dict[str, int] = {}
        current_section: str | None = None
        seq_in_section = 0
        prev_track_row: int | None = None
        section_banner_count = 0
        first_section_banner_title: str | None = None
        section_sort_state = section_sort.SectionSortState(
            track=section_sort.default_track(sheet_name),
            scope=None,
        )
        sub_chain_col = SUB_CHAIN_BOUNDARY_COLUMN.get(sheet_name)
        prev_sub_chain_value: str | None = None

        for r_idx in range(2, (ws.max_row or 0) + 1):
            a_cell = ws.cell(row=r_idx, column=1)
            a_val = norm_value(a_cell.value)
            a_fill = cell_fill(a_cell)

            data: dict[str, str] = {}
            for c in columns:
                cell = ws.cell(row=r_idx, column=c["index"])
                v = norm_value(cell.value)
                if v is not None:
                    data[c["key"]] = v

            is_banner = a_fill == SECTION_FILL or (
                a_val is not None and not data and a_val.isupper() and len(a_val) >= 4
            )

            if is_banner:
                section_banner_count += 1
                if first_section_banner_title is None:
                    first_section_banner_title = (a_val or sheet_name).title()
                current_section = (a_val or "").title() or sheet_name
                seq_in_section = 0
                prev_track_row = None
                prev_sub_chain_value = None
                section_payload: dict[str, object] = dict(data)
                if section_sort.supports_sheet(sheet_name):
                    section_payload["section_sort"] = section_sort.classify_section(
                        sheet_name,
                        a_val or "",
                        r_idx,
                        section_sort_state,
                    )
                node_rows.append(
                    (
                        run_id,
                        sheet_name,
                        r_idx,
                        a_val,
                        "todo",
                        "section",
                        current_section,
                        0,
                        json.dumps(section_payload),
                        None,
                    )
                )
                continue

            if is_inline_section_marker(a_val, data):
                current_section = a_val
                seq_in_section = 0
                prev_track_row = None
                prev_sub_chain_value = None
                node_rows.append(
                    (
                        run_id,
                        sheet_name,
                        r_idx,
                        a_val,
                        "todo",
                        "section",
                        current_section,
                        0,
                        json.dumps({}),
                        None,
                    )
                )
                continue

            if not data and a_val is None:
                continue

            label = data.get(label_key) if label_key else None
            if not label:
                label = next(iter(data.values()), None)

            state = parse_state(a_val)
            if a_val is None and not is_value_sheet:
                state = "todo"

            row_type = "value" if is_value_sheet else "checkbox"

            node_rows.append(
                (
                    run_id,
                    sheet_name,
                    r_idx,
                    label,
                    state,
                    row_type,
                    current_section,
                    seq_in_section,
                    json.dumps(data),
                    _row_hash(data),
                )
            )
            if label:
                label_to_row.setdefault(label.strip().lower(), r_idx)

            # Only emit prerequisite edges inside *real* chain sections.
            # Sidequest collections / FATEs / crafting logs share section
            # context but their rows are independently completable, so they
            # produce no sequence edges and won't trigger cascades.
            in_chain = section_is_chain(sheet_name, current_section)
            if sub_chain_col is not None:
                cur_sub = data.get(sub_chain_col)
                if (
                    prev_sub_chain_value is not None
                    and cur_sub != prev_sub_chain_value
                ):
                    prev_track_row = None
                prev_sub_chain_value = cur_sub
            if row_type == "checkbox" and prev_track_row is not None and in_chain:
                seq_edges.append(
                    (
                        run_id,
                        sheet_name,
                        "sequence",
                        prev_track_row,
                        None,
                        r_idx,
                        label,
                        1,
                    )
                )
            prev_track_row = r_idx
            seq_in_section += 1

            if unlock_col and unlock_col["key"] in data:
                for cand in split_candidates(data[unlock_col["key"]]):
                    unlock_edges.append(
                        (
                            run_id,
                            sheet_name,
                            "unlocks",
                            r_idx,
                            label,
                            None,
                            cand,
                            0,
                        )
                    )

        resolved_unlocks = []
        for e in unlock_edges:
            tgt = label_to_row.get((e[6] or "").strip().lower())
            resolved_unlocks.append(e[:5] + (tgt, e[6], 1 if tgt else 0))

        track_total = sum(1 for n in node_rows if n[5] in ("checkbox", "value"))
        # If a sheet has multiple section banners, row 2's banner is just the
        # first in-sheet section (e.g., a zone) rather than a true sheet name.
        # Keep the workbook sheet name to avoid duplicate card titles.
        sheet_title = (
            sheet_name
            if section_banner_count >= 2
            else (first_section_banner_title or sheet_name)
        )

        conn.execute(
            """
            INSERT INTO sheets (
                run_id, sheet_index, sheet_name, title,
                is_menu, is_readonly, parent_sheet, data_columns_json,
                label_key, value_key, total_rows
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                sheet_index,
                sheet_name,
                sheet_title,
                0,
                0,
                parent,
                json.dumps(columns),
                label_key,
                value_key,
                track_total,
            ),
        )

        conn.executemany(
            """
            INSERT INTO nodes (
                run_id, sheet_name, row_index, label,
                baseline_state, row_type, section_label, seq, row_json,
                stable_hash
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            node_rows,
        )
        log_memory_checkpoint(
            mem_log,
            f"sheet {sheet_index}/{sheet_count}: nodes inserted {sheet_name} count={len(node_rows)}",
        )

        conn.executemany(
            """
            INSERT INTO edges (
                run_id, sheet_name, edge_type,
                source_row_index, source_label, target_row_index,
                target_label, resolved
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            seq_edges + resolved_unlocks,
        )
        log_memory_checkpoint(
            mem_log,
            f"sheet {sheet_index}/{sheet_count}: edges inserted {sheet_name} count={len(seq_edges) + len(resolved_unlocks)}",
        )

        total_rows += track_total
        print(
            f"  [{sheet_index:3}] {sheet_name:<52} rows={track_total:<5} "
            f"edges={len(seq_edges) + len(resolved_unlocks)} parent={parent}"
        )
        log_memory_checkpoint(
            mem_log,
            f"sheet {sheet_index}/{sheet_count}: done {sheet_name}",
        )

    if saved_chars:
        conn.executemany(
            "INSERT INTO characters (id, name, starting_class, created_at) VALUES (?, ?, ?, ?)",
            [(c[0], c[1], None, c[2]) if len(c) == 3 else c for c in saved_chars],
        )
    else:
        conn.execute(
            "INSERT INTO characters (name, starting_class, created_at) VALUES (?, ?, ?)",
            ("Adventurer", None, started),
        )

    if saved_progress:
        conn.executemany(
            """
            INSERT OR IGNORE INTO character_progress
                (character_id, run_id, sheet_name, row_index, state, progress_percent, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [(p[0], run_id, p[1], p[2], p[3], p[4], p[5]) for p in saved_progress],
        )
        print(f"\nMigrated {len(saved_progress)} progress rows to run {run_id}")

    # --- Second pass: repair child parentage + capture menu sections ------
    # Some workbook tabs link "Main Page" back to a top menu even though they
    # are visually nested under a submenu. Use menu-sheet hyperlinks as the
    # structural source of truth so future ingests keep submenu children in the
    # correct branch.
    print("\nRepairing child parent links from menu hyperlinks...")
    log_memory_checkpoint(mem_log, "ingest: repair pass start")

    menu_rows = list(conn.execute(
        "SELECT sheet_name, sheet_index FROM sheets "
        "WHERE run_id = ? AND is_menu = 1 ORDER BY sheet_index",
        (run_id,),
    ))
    content_sheet_names = {
        r[0] for r in conn.execute(
            "SELECT sheet_name FROM sheets "
            "WHERE run_id = ? AND is_menu = 0 AND is_readonly = 0",
            (run_id,),
        )
    }

    # Count how many sibling submenus under each top menu link to a target.
    # If a target appears in multiple submenus, treat it as a shared cross-link
    # and keep it under the top menu instead of arbitrarily nesting it under
    # whichever submenu happens to run first.
    submenu_target_counts: dict[str, dict[str, int]] = {}
    for menu_row in menu_rows:
        menu_name = menu_row[0]
        top_parent = menu_parent(menu_name)
        if top_parent is None:
            continue
        try:
            ws = wb[menu_name]
        except KeyError:
            continue
        targets = {
            t
            for t in extract_menu_link_targets(ws)
            if t in content_sheet_names
        }
        if not targets:
            continue
        counts = submenu_target_counts.setdefault(top_parent, {})
        for target in targets:
            counts[target] = counts.get(target, 0) + 1

    relinked_total = 0
    for menu_row in menu_rows:
        menu_name = menu_row[0]
        try:
            ws = wb[menu_name]
        except KeyError:
            continue

        targets = sorted(t for t in extract_menu_link_targets(ws) if t in content_sheet_names)
        if not targets:
            continue

        submenu_parent = menu_parent(menu_name)
        applied = 0
        for target in targets:
            if submenu_parent is not None:
                target_links = submenu_target_counts.get(submenu_parent, {})
                if target_links.get(target, 0) != 1:
                    # Shared submenu cross-link: leave this sheet at the
                    # top-menu level.
                    continue

            if submenu_parent is None:
                # Top menus should only adopt currently-unparented sheets.
                cur = conn.execute(
                    "UPDATE sheets SET parent_sheet = ? "
                    "WHERE run_id = ? AND sheet_name = ? "
                    "AND is_menu = 0 AND is_readonly = 0 "
                    "AND parent_sheet IS NULL",
                    (menu_name, run_id, target),
                )
            else:
                # Submenus may override a top-menu parent from Main Page links.
                cur = conn.execute(
                    "UPDATE sheets SET parent_sheet = ? "
                    "WHERE run_id = ? AND sheet_name = ? "
                    "AND is_menu = 0 AND is_readonly = 0 "
                    "AND (parent_sheet IS NULL OR parent_sheet = ? OR parent_sheet = ?)",
                    (menu_name, run_id, target, submenu_parent, menu_name),
                )
            applied += int(cur.rowcount or 0)

        relinked_total += applied
        if applied:
            print(f"  [{menu_name}]: re-parented {applied} child sheets")

    print(f"  total re-parented: {relinked_total}")

    override_total = 0
    for child_sheet, parent_sheet in CONTENT_PARENT_OVERRIDES.items():
        if child_sheet not in content_sheet_names or parent_sheet not in content_sheet_names:
            continue
        cur = conn.execute(
            "UPDATE sheets SET parent_sheet = ? "
            "WHERE run_id = ? AND sheet_name = ? "
            "AND is_menu = 0 AND is_readonly = 0",
            (parent_sheet, run_id, child_sheet),
        )
        override_total += int(cur.rowcount or 0)
    if override_total:
        print(f"  explicit parent overrides applied: {override_total}")

    log_memory_checkpoint(mem_log, "ingest: repair pass complete")

    # --- Third pass: capture menu-sheet column groupings ------------------
    # Each menu sheet's children are physically arranged in side-by-side
    # columns under group headers (e.g. Duty Menu - Journal: Main Scenario /
    # Sidequests / Allied Society / Other Quests). The first ingest pass
    # only recorded each child's parent menu — this pass figures out which
    # *column* under that menu it belongs to, by walking the menu sheet's
    # layout and matching child-sheet names to entries in each column.
    print("\nMatching child sheets to menu sections...")
    log_memory_checkpoint(mem_log, "ingest: menu-section match pass start")

    # Hand-curated aliases for menu labels that intentionally differ from
    # their sheet name (display polish in the workbook).
    LABEL_ALIASES = {
        "hildibrand sidequests": "hildibrand",
        "weapon enhancement sidequests": "weapon enhancement",
        "seasonal events": "seasonal quests",
    }

    def _norm_match(text: str) -> str:
        """Aggressive normalization for menu-label vs sheet-name matching.

        * strip parenthetical disambiguators (``"Main Scenario (Dawntrail)"``)
        * strip apostrophes (``Amalj'aa`` ↔ ``Amaljaa``, ``Ul'dahn`` ↔ ``Uldahn``)
        * strip ``main scenario`` (menu has the wider label, sheets don't)
        * strip ``chronicles of a new era - `` prefix so e.g. that menu's
          ``Chronicles of a New Era - Bahamut`` resolves to sheet ``Bahamut``
        """
        t = text.lower().strip()
        t = re.sub(r"\s*\([^)]*\)\s*", " ", t)
        t = t.replace("'", "")
        t = re.sub(r"^chronicles of a new era\s*[- ]\s*", "", t)
        for noise in (" main scenario ", " main scenario"):
            t = t.replace(noise, " ")
        return " ".join(LABEL_ALIASES.get(t.strip(), t).split())

    def _strip_quests_suffix(t: str) -> str:
        """Drop a trailing ' quests' so menu labels like 'Yok Huy Quests'
        match sheet names like 'Yok Huy'. Symmetric — applied to both sides."""
        return re.sub(r"\s+quests$", "", t)

    sheets_lower: dict[str, str] = {}
    sheets_norm: dict[str, str] = {}
    sheets_loose: dict[str, str] = {}  # last-resort: also strip trailing " quests"
    for r in conn.execute(
        "SELECT sheet_name FROM sheets WHERE run_id = ? AND is_menu = 0",
        (run_id,),
    ):
        name = r[0]
        sheets_lower.setdefault(name.lower(), name)
        sheets_norm.setdefault(_norm_match(name), name)
        sheets_loose.setdefault(_strip_quests_suffix(_norm_match(name)), name)

    def _resolve_label(label: str) -> str | None:
        """Try every alias / normalization tier; return the matching sheet
        name or None. Shared between the parser (for sub-header detection)
        and the post-parse update loop (for actually applying sections)."""
        return (
            sheets_lower.get(label.lower())
            or sheets_norm.get(_norm_match(label))
            or sheets_loose.get(_strip_quests_suffix(_norm_match(label)))
        )

    grouped_total = 0
    unmatched_total = 0
    unmatched_samples: list[str] = []
    for menu_name in [r[0] for r in menu_rows]:
        try:
            ws = wb[menu_name]
        except KeyError:
            continue
        # Pass _resolve_label so the parser can distinguish real child rows
        # from in-column sub-headers ("Chronicles of a New Era", etc.).
        groups = extract_menu_groups(
            ws, is_real_sheet=lambda t: _resolve_label(t) is not None,
        )
        if not groups:
            continue
        applied = 0
        for child_lower, group_label in groups.items():
            real_name = _resolve_label(child_lower)
            if real_name is None:
                unmatched_total += 1
                if len(unmatched_samples) < 8:
                    unmatched_samples.append(f"[{menu_name}] {child_lower!r}")
                continue
            cur = conn.execute(
                "UPDATE sheets SET parent_menu_section = ? "
                "WHERE run_id = ? AND sheet_name = ? AND parent_sheet = ?",
                (group_label, run_id, real_name, menu_name),
            )
            if cur.rowcount > 0:
                applied += 1
        grouped_total += applied
        print(f"  [{menu_name}]: {applied} children grouped into {len(set(groups.values()))} sections")
    print(f"  total grouped: {grouped_total}")
    if unmatched_total:
        print(f"  unmatched menu entries: {unmatched_total} "
              "(likely group/sub-headers without a real sheet)")
        for sample in unmatched_samples:
            print(f"    {sample}")
    log_memory_checkpoint(mem_log, "ingest: menu-section match pass complete")

    # Release the data-only workbook before loading the formula workbook.
    # Keeping both loaded at once can cause very high peak RAM on large files.
    try:
        wb.close()
    except Exception:
        pass
    log_memory_checkpoint(mem_log, "ingest: data workbook closed")
    del wb
    gc.collect()
    log_memory_checkpoint(mem_log, "ingest: gc after data workbook close")

    print("\nScanning class-conditional formulas (second workbook pass)...")
    overrides = collect_class_overrides(xlsx_path, run_id, mem_log=mem_log)
    if overrides:
        conn.executemany(
            """
            INSERT INTO class_overrides
                (run_id, starting_class, sheet_name, row_index, state)
            VALUES (?, ?, ?, ?, ?)
            """,
            overrides,
        )
        affected_cells = len({(s, r) for _, _, s, r, _ in overrides})
        per_sheet: dict[str, int] = {}
        for _, _, s, _, _ in overrides:
            per_sheet[s] = per_sheet.get(s, 0) + 1
        print(f"  class_overrides: {len(overrides)} rows across {affected_cells} cells")
        for s, n in sorted(per_sheet.items(), key=lambda x: -x[1]):
            print(f"    {s:<52} {n}")

    completed = dt.datetime.now().isoformat(timespec="seconds")
    conn.execute(
        "UPDATE ingest_runs SET completed_at = ?, sheet_count = ?, row_count = ? WHERE id = ?",
        (completed, sheet_count, total_rows, run_id),
    )

    conn.commit()
    log_memory_checkpoint(mem_log, "ingest: final commit complete")
    conn.close()
    log_memory_checkpoint(mem_log, "ingest: connection closed")

    print(f"\nIngest complete -> run {run_id}")
    print(f"  sheets : {sheet_count}")
    print(f"  rows   : {total_rows}")


def resolve_xlsx(explicit: Path | None) -> Path:
    if explicit:
        return explicit

    # Default source is the Spreadsheet folder, regardless of workbook filename.
    spreadsheet_dir: Path | None = None
    for child in Path.cwd().iterdir():
        if child.is_dir() and child.name.lower() == "spreadsheet":
            spreadsheet_dir = child
            break

    if spreadsheet_dir is None:
        raise SystemExit("Spreadsheet folder not found in project root.")

    candidates = sorted(
        (p for p in spreadsheet_dir.glob("*.xlsx") if not p.name.startswith("~$")),
        key=lambda p: p.stat().st_mtime,
    )
    if not candidates:
        raise SystemExit(f"No .xlsx found in {spreadsheet_dir}; pass --xlsx <path>.")
    print(f"Auto-selected workbook: {candidates[-1]}")
    return candidates[-1]


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest the FFXIV workbook into SQLite.")
    parser.add_argument(
        "--xlsx",
        type=Path,
        help="source workbook override (default: newest .xlsx in Spreadsheet folder)",
    )
    parser.add_argument("--db", type=Path, default=Path("data/ffxiv_tracker.sqlite"))
    parser.add_argument(
        "--mem-log",
        action="store_true",
        help="print process memory checkpoints during ingest",
    )
    args = parser.parse_args()

    ingest(resolve_xlsx(args.xlsx), args.db, mem_log=args.mem_log)


if __name__ == "__main__":
    main()
